import os
import pandas as pd
from datetime import date, timedelta
from openpyxl import load_workbook
import shutil


class Date_Stamps():

    the_date = None

    def __init__(self):
        self.time = date.today()

    def get_today(self):
        return self.time

    def get_yesterday(self):
        the_date = self.time - timedelta(days=1)
        return the_date

    def count_in_weekend(self):
        if self.time.isoweekday() == 1:
            the_date = self.time - timedelta(days=3)
            return the_date
        else:
            the_date = self.time - timedelta(days=1)
            return the_date

    def get_start_month(self):
        the_date = self.time.replace(day=1)
        return the_date

    def __str__(self):
        return f'today is {self.time}'


current_date = Date_Stamps().get_today()

dst = f'/home/mike/CSAT_Reports/{current_date} reports'
templates_repo = ['/home/mike/github/ReportAutomation/Reports/ReportTemplates/daily_template.xlsx',
                  '/home/mike/github/ReportAutomation/Reports/ReportTemplates/report_maker.xlsx']


def main_folder_creation():
    if not os.path.exists('/home/mike/CSAT_Reports'):
        os.mkdir('/home/mike/CSAT_Reports')


def folder_creation(item):
    global current_date
    if not os.path.exists(item):
        os.mkdir(item)


def copy_template(template, destination):
    for item in template:
        shutil.copy2(item, destination)
    os.rename(
        f'{dst}/daily_template.xlsx',
        f'{dst}/CSAT Daily HS {current_date}.xlsx', )
    os.rename(
        f'{dst}/report_maker.xlsx',
        f'{dst}/report_maker {current_date}.xlsx', )


def write_daily_report(data):
    book = load_workbook(f'{dst}/CSAT Daily HS {current_date}.xlsx')
    writer = pd.ExcelWriter(f'{dst}/CSAT Daily HS {current_date}.xlsx',
                            engine='openpyxl', mode='a')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, sheet_name='Negative Comments', startrow=1, index=False, header=False)
    writer.save()


def write_daily_report_maker(data, data_1):
    book = load_workbook(f'{dst}/report_maker {current_date}.xlsx')
    writer = pd.ExcelWriter(f'{dst}/report_maker {current_date}.xlsx', engine='openpyxl', mode='a')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, sheet_name='Formulas&Pivots', startrow=1, index=False, header=False)
    data_1.to_excel(writer, sheet_name='Formulas&Pivots', startrow=1,
                    startcol=7, index=False, header=False)
    writer.save()
